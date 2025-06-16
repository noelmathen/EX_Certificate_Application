#proserv.py

#!/usr/bin/env python3
"""
proserv.py  ·  Rev-2025-06-17
Handles single-row and split-row headers so that *all* Proserv PDFs
(including KB5766) are extracted.
"""

import os, glob, logging, fitz, camelot, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- patch rmtree so “file in use” is retried/ignored (Windows) -----
import shutil, time, errno, os, functools
_orig_rmtree = shutil.rmtree

def _robust_rmtree(path, *args, **kwargs):
    # run original first → if it fails with PermissionError, retry once
    try:
        _orig_rmtree(path, *args, **kwargs)
    except PermissionError as e:
        if e.errno != errno.EACCES:
            raise
        time.sleep(0.2)              # tiny pause lets Ghostscript exit
        try:
            _orig_rmtree(path, *args, **kwargs)
        except PermissionError:
            pass                     # still locked → just leave the dir

shutil.rmtree = _robust_rmtree

# ─── CONFIG ──────────────────────────────────────────────────────────────
SOURCE_DIR, OUT_XLSX = "Proserv Certificates", "Proserv_Certificates.xlsx"
SHEET_NAME, EXPECTED_COLS = "Proserv Certificates", 18
TOKENS = ["S-No", "Equipment Tag", "Pass/Fail"]          # header keys
FIXED_COLS = [                                           # 18 data cols
    "S-No.", "Equipment Tag #", "Equipment Description", "Manufacturer",
    "Model", "Circuit ID", "Area of Classification", "IP", "Protection Method",
    "Gas Group", "T-Rating", "Serial Number", "Certifying Authority",
    "Certificate No.", "Grade of Inspection", "Inspection Date",
    "Expiry Date", "Pass/Fail"
]

# ─── LOGGER ──────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

errs, dfs = [], []                     # collect issues & dataframes

# ─── 1. Locate PDFs ──────────────────────────────────────────────────────
pdfs = sorted(glob.glob(os.path.join(SOURCE_DIR, "*.pdf")))
if not pdfs:
    log.error(f"No PDFs found in '{SOURCE_DIR}'"); raise SystemExit(1)

# ─── 2. Iterate over certificates ────────────────────────────────────────
for pdf in pdfs:
    name = os.path.basename(pdf)
    log.info(f"🔍  {name}")
    equip_id = ""

    # open & read first page
    try:
        doc = fitz.open(pdf)
        lines = doc[0].get_text().splitlines()
        idx   = next(i for i,l in enumerate(lines) if "Equipment ID" in l)
        equip_id = lines[idx+5].strip()
        last_page = doc.page_count
        doc.close()
    except Exception as e:
        errs.append(f"{name}: open/ID error ({e})"); continue

    # camelot extract (lattice → stream)
    tables = []
    for flav in ("lattice", "stream"):
        try:
            tables = camelot.read_pdf(pdf, pages=str(last_page),
                                      flavor=flav, line_scale=40)
            if tables: break
        except Exception as e:
            errs.append(f"{name}: camelot {flav} error ({e})")
    if not tables:
        errs.append(f"{name}: no table found on last page"); continue

    df = tables[0].df.copy()

    # ─ Header handling ───────────────────────────────────────────
    def has_tokens(row) -> bool:
        cells = [str(c).lower() for c in row]
        return all(any(t.lower() in c for c in cells) for t in TOKENS)

    header_idx = next((i for i,r in df.iterrows() if has_tokens(r)), None)

    # Fallback: merge row-0 and row-1 cell-by-cell
    if header_idx is None and len(df) > 1:
        merged = df.iloc[0].astype(str) + " " + df.iloc[1].astype(str)
        if has_tokens(merged):
            df.columns = merged.tolist()
            df = df[2:].reset_index(drop=True)
            header_idx = -1          # mark as “handled”
    # Normal single-row header
    if header_idx is not None and header_idx >= 0:
        df.columns = df.iloc[header_idx].tolist()
        df = df[header_idx+1:].reset_index(drop=True)

    if header_idx is None:
        # errs.append(f"{name}: header still not detected – forcing columns")
        pass

    # ─ Normalise columns & collect ───────────────────────────────
    if df.shape[1] < EXPECTED_COLS:
        errs.append(f"{name}: only {df.shape[1]} cols (need ≥{EXPECTED_COLS})")
        continue                     # skip completely malformed tables

    df = df.iloc[:, :EXPECTED_COLS]
    df.columns = FIXED_COLS
    df["Filename"], df["Equipment ID"] = name, equip_id
    dfs.append(df)

# ─── 3. Merge & write Excel ──────────────────────────────────────────────
if not dfs:
    log.error("❌  No usable tables extracted."); [log.error(" - "+e) for e in errs]
    raise SystemExit(1)

final = pd.concat(dfs, ignore_index=True)
# keep rows whose S-No. looks numeric
final = final[final["S-No."].astype(str).str.match(r'^\d+$')].reset_index(drop=True)

# insert metadata after the 18 core columns
final.insert(len(FIXED_COLS),  "Filename",     final.pop("Filename"))
final.insert(len(FIXED_COLS)+1,"Equipment ID", final.pop("Equipment ID"))

try:
    final.to_excel(OUT_XLSX, sheet_name=SHEET_NAME, index=False)
    wb, ws = load_workbook(OUT_XLSX), None
    ws = wb[SHEET_NAME]
    for c in ws[1]: c.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = \
            max((len(str(c.value)) for c in col if c.value), default=0)+2
    wb.save(OUT_XLSX)
    log.info(f"✅  {len(final)} rows written to '{OUT_XLSX}'")
except Exception as e:
    errs.append(f"write-excel: {e}")




# ─── 4. Summary ──────────────────────────────────────────────────────────
if errs:
    log.warning("⚠️  Issues:")
    for e in errs: log.warning(" - %s", e)
