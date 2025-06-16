#src/oman_processor.py
#!/usr/bin/env python3
"""
Oman Certificate Processor Module
Handles processing of Oman certificates into Excel format
"""

import os
import re
import glob
import logging
import fitz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

class OmanProcessor:
    """Oman certificate processor"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Configuration
        self.SOURCE_DIR = "Oman Certificates"
        self.OUT_XLSX = "Oman_Certificates.xlsx"
        self.SHEET_NAME = "Oman Certificates"
        self.Y_TOLERANCE = 5  # Vertical tolerance for span matching
        
        # Field definitions (27 data fields + Filename)
        self.FIELDS = [
            "PROJECT DESCRIPTION", "CLIENT TAG", "CIRCUIT ID", "DESCRIPTION", "SYSTEM",
            "MANUFACTURER", "TYPE/MODEL", "SERIAL NUMBER", "Ex PROTECTION", "EPL",
            "CERTIFIED BODY", "Ex CERT No", "DATE INSPECTED", "PROJECT WBS NO",
            "EX INSPECTION TAG No", "AREA CLASSIFICATION", "AREA CLASS", "LAYOUT DWG",
            "LOCATION", "AREA", "GRID REFERENCE", "ACCESS ARRANGEMENT", "IP RATING",
            "REPAIR CATEGORY", "NEXT INSPECTION DUE", "PASS / FAIL",
            "Filename"
        ]
        
        # Create normalized key mapping
        self.normalized_map = {self._normalize_key(field): field 
                              for field in self.FIELDS if field != "Filename"}
        
        # Storage for results
        self.records = []
        self.errors = []
        
    def _normalize_key(self, text):
        """Normalize text for key matching by removing non-alphanumeric characters"""
        return re.sub(r'[^A-Z0-9]', "", text.upper())
        
    def process(self):
        """Main processing function"""
        self.logger.info("Starting Oman certificate processing...")
        
        # Find PDF files
        pdf_files = sorted(glob.glob(os.path.join(self.SOURCE_DIR, "*.pdf")))
        if not pdf_files:
            self.logger.error(f"No PDFs found in '{self.SOURCE_DIR}'")
            return
            
        self.logger.info(f"Found {len(pdf_files)} Oman certificates to process")
        
        # Process each PDF
        for idx, pdf_file in enumerate(pdf_files, 1):
            self._process_single_pdf(pdf_file, idx)
            
        # Generate Excel file
        self._generate_excel()
        
        # Log summary
        self._log_summary()
        
    def _process_single_pdf(self, pdf_path, sequence_number):
        """Process a single PDF file"""
        filename = os.path.basename(pdf_path)
        self.logger.info(f"🔍 ({sequence_number}/{len(glob.glob(os.path.join(self.SOURCE_DIR, '*.pdf')))}) {filename}")
        
        try:
            # Extract and parse certificate data
            metadata, missing_fields = self._parse_certificate(pdf_path, sequence_number)
            
            if missing_fields:
                # Create blank record for manual fixing
                self.errors.append(f"{filename}: missing {len(missing_fields)} fields → {missing_fields}")
                blank_record = {"Sl. No": sequence_number, "Filename": filename}
                blank_record.update({field: "" for field in self.FIELDS})
                self.records.append(blank_record)
            else:
                self.records.append(metadata)
                self.logger.info(f"✅ Successfully processed: {filename}")
                
        except Exception as e:
            self.errors.append(f"{filename}: Exception - {e}")
            self.logger.error(f"❌ Error processing {filename}: {e}")
            
            # Add blank record
            blank_record = {"Sl. No": sequence_number, "Filename": filename}
            blank_record.update({field: "" for field in self.FIELDS})
            self.records.append(blank_record)
            
    def _parse_certificate(self, pdf_path, sequence_number):
        """Parse certificate data from PDF"""
        # Extract text spans from first page
        spans = self._extract_spans(pdf_path)
        
        # Extend spans with merged field keys
        spans = self._extend_spans_with_merged(spans)
        
        # Separate keys and values
        keys = [span for span in spans if span["normalized_key"] in self.normalized_map]
        values = [span for span in spans if span["normalized_key"] not in self.normalized_map]
        
        # Initialize metadata
        metadata = {
            "Sl. No": sequence_number,
            "Filename": os.path.basename(pdf_path)
        }
        
        # Match each key with its nearest value
        for key_span in keys:
            field_name = self.normalized_map[key_span["normalized_key"]]
            x_end = key_span["bbox"][2]
            y_mid = key_span["y_position"]
            
            # Find candidate values on the same line to the right of the key
            candidates = []
            for value_span in values:
                if (abs(value_span["y_position"] - y_mid) <= self.Y_TOLERANCE and 
                    value_span["bbox"][0] > x_end):
                    distance = value_span["bbox"][0] - x_end
                    candidates.append((distance, value_span))
                    
            # Use the closest value
            if candidates:
                closest_value = min(candidates)[1]["text"].strip()
                metadata[field_name] = closest_value
            else:
                metadata[field_name] = ""
                
        # Ensure all fields are present
        for field in self.FIELDS:
            metadata.setdefault(field, "")
            
        # Find missing fields
        missing_fields = [field for field in self.FIELDS if not metadata[field].strip()]
        
        return metadata, missing_fields
        
    def _extract_spans(self, pdf_path):
        """Extract text spans from the first page of the PDF"""
        spans = []
        
        with fitz.open(pdf_path) as doc:
            page_data = doc[0].get_text("dict")
            
            for block in page_data["blocks"]:
                if block.get("type") != 0:  # Skip non-text blocks
                    continue
                    
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"].strip()
                        if not text:
                            continue
                            
                        x0, y0, x1, y1 = span["bbox"]
                        spans.append({
                            "text": text,
                            "normalized_key": self._normalize_key(text),
                            "bbox": (x0, y0, x1, y1),
                            "y_position": (y0 + y1) / 2
                        })
                        
        return spans
        
    def _extend_spans_with_merged(self, spans):
        """Extend spans with merged field keys for broken text"""
        # Group spans by row (using Y tolerance)
        rows = {}
        for span in spans:
            row_key = int(span["y_position"] // self.Y_TOLERANCE)
            rows.setdefault(row_key, []).append(span)
            
        merged_spans = []
        
        # For each row, try concatenating adjacent spans
        for row_spans in rows.values():
            row_spans.sort(key=lambda s: s["bbox"][0])  # Sort left to right
            
            for i in range(len(row_spans)):
                for j in range(i + 1, min(i + 5, len(row_spans))):  # Try up to 4 more spans
                    # Concatenate text from span i to span j
                    combined_text = "".join(span["text"] for span in row_spans[i:j+1])
                    normalized_combined = self._normalize_key(combined_text)
                    
                    if normalized_combined in self.normalized_map:
                        # Create merged span
                        x0 = row_spans[i]["bbox"][0]
                        y0 = row_spans[i]["bbox"][1]
                        x1 = row_spans[j]["bbox"][2]
                        y1 = row_spans[j]["bbox"][3]
                        
                        merged_spans.append({
                            "text": combined_text,
                            "normalized_key": normalized_combined,
                            "bbox": (x0, y0, x1, y1),
                            "y_position": (row_spans[i]["y_position"] + row_spans[j]["y_position"]) / 2
                        })
                        break
                        
        return spans + merged_spans
        
    def _generate_excel(self):
        """Generate the Excel output file"""
        if not self.records:
            self.logger.error("No records to write to Excel")
            return
            
        try:
            # Create DataFrame
            df = pd.DataFrame(self.records, columns=["Sl. No"] + self.FIELDS)
            
            # Write to Excel
            df.to_excel(self.OUT_XLSX, sheet_name=self.SHEET_NAME, index=False)
            
            # Format Excel file
            self._format_excel_file()
            
            self.logger.info(f"✅ Wrote {len(self.records)} rows to '{self.OUT_XLSX}'")
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel file: {e}")
            raise
            
    def _format_excel_file(self):
        """Format the Excel file with proper styling"""
        try:
            wb = load_workbook(self.OUT_XLSX)
            ws = wb[self.SHEET_NAME]
            
            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = max(
                    len(str(cell.value)) for cell in column if cell.value
                ) + 2
                ws.column_dimensions[column[0].column_letter].width = max_length
                
            wb.save(self.OUT_XLSX)
            
        except Exception as e:
            self.logger.warning(f"Could not format Excel file: {e}")
            
    def _log_summary(self):
        """Log processing summary"""
        total_processed = len(self.records)
        total_errors = len(self.errors)
        
        self.logger.info("📊 Oman Processing Summary:")
        self.logger.info(f"   • Successfully processed: {total_processed}")
        self.logger.info(f"   • Records with issues: {total_errors}")
        
        if self.errors:
            self.logger.warning("⚠️ Processing issues:")
            for error in self.errors:
                self.logger.warning(f" - {error}")