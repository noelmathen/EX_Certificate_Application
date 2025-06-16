#!/usr/bin/env python3
import os
import re
import glob
import logging

import fitz             # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# —————————————————————————————————————
# CONFIGURATION
# —————————————————————————————————————

SOURCE_DIR = "Oman Certificates"
OUT_XLSX   = "Oman_Certificates.xlsx"
SHEET_NAME = "Oman Certificates"

# 27 fields + our new 'Filename'
FIELDS = [
    "PROJECT DESCRIPTION", "CLIENT TAG", "CIRCUIT ID", "DESCRIPTION", "SYSTEM",
    "MANUFACTURER", "TYPE/MODEL", "SERIAL NUMBER", "Ex PROTECTION", "EPL",
    "CERTIFIED BODY", "Ex CERT No", "DATE INSPECTED", "PROJECT WBS NO",
    "EX INSPECTION TAG No", "AREA CLASSIFICATION", "AREA CLASS", "LAYOUT DWG",
    "LOCATION", "AREA", "GRID REFERENCE", "ACCESS ARRANGEMENT", "IP RATING",
    "REPAIR CATEGORY", "NEXT INSPECTION DUE", "PASS / FAIL",
    "Filename"
]

Y_TOL = 5  # pts tolerance when matching spans by Y

def normalize_key(s: str) -> str:
    return re.sub(r'[^A-Z0-9]', "", s.upper())

# Build lookup normalized→field (exclude Filename)
NORMALIZED_MAP = { normalize_key(f): f for f in FIELDS if f != "Filename" }

# —————————————————————————————————————
# LOGGER SETUP
# —————————————————————————————————————

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

# —————————————————————————————————————
# 1) EXTRACT ALL SPANS FROM PAGE 1
# —————————————————————————————————————

def extract_spans(pdf_path):
    doc  = fitz.open(pdf_path)
    page = doc[0]
    data = page.get_text("dict")
    spans = []
    for block in data["blocks"]:
        if block.get("type") != 0:  # text only
            continue
        for line in block["lines"]:
            for sp in line["spans"]:
                txt = sp["text"].strip()
                if not txt:
                    continue
                x0, y0, x1, y1 = sp["bbox"]
                spans.append({
                    "text": txt,
                    "kn":    normalize_key(txt),
                    "bbox":  (x0, y0, x1, y1),
                    "yp":    (y0 + y1) / 2
                })
    return spans

# —————————————————————————————————————
# 2) MERGE BROKEN KEYS (like "PROJECT DESC"+"R"+"IPTION")
# —————————————————————————————————————

def extend_spans_with_merged(spans):
    # bucket spans into horizontal rows
    rows = {}
    for s in spans:
        key = int(s["yp"] // Y_TOL)
        rows.setdefault(key, []).append(s)
    merged = []
    for row in rows.values():
        row = sorted(row, key=lambda s: s["bbox"][0])
        n = len(row)
        for i in range(n):
            text = row[i]["text"]
            x0,y0,_,_ = row[i]["bbox"]
            # try concatenating up to 4 subsequent spans
            for j in range(i+1, min(i+5, n)):
                text += row[j]["text"]
                kn = normalize_key(text)
                if kn in NORMALIZED_MAP:
                    x1 = row[j]["bbox"][2]
                    merged.append({
                        "text": text,
                        "kn":    kn,
                        "bbox": (x0, y0, x1, row[j]["bbox"][3]),
                        "yp":   (row[i]["yp"] + row[j]["yp"]) / 2
                    })
                    break
    return spans + merged

# —————————————————————————————————————
# 3) PARSE ONE PDF → (meta, missing_fields)
# —————————————————————————————————————

def parse_certificate(pdf_path, seq_no):
    spans = extract_spans(pdf_path)
    spans = extend_spans_with_merged(spans)

    # split into keys vs values
    key_spans, value_spans = [], []
    for s in spans:
        if s["kn"] in NORMALIZED_MAP:
            key_spans.append({
                "field": NORMALIZED_MAP[s["kn"]],
                "bbox":  s["bbox"],
                "yp":    s["yp"]
            })
        else:
            value_spans.append(dict(s))

    # build metadata dict
    meta = {"Sl. No": seq_no, "Filename": os.path.basename(pdf_path)}

    # pair each key with nearest right-hand value on same line
    for k in key_spans:
        fld = k["field"]
        x_end = k["bbox"][2]
        y_mid = k["yp"]
        # candidates: same y band, to right of key
        cands = [
            (vs["bbox"][0] - x_end, vs)
            for vs in value_spans
            if abs(vs["yp"] - y_mid) <= Y_TOL and vs["bbox"][0] > x_end
        ]
        if not cands:
            meta[fld] = ""
        else:
            cands.sort(key=lambda x: x[0])
            chosen = cands[0][1]
            meta[fld] = chosen["text"].strip()
            value_spans.remove(chosen)

    # ensure every field exists
    for f in FIELDS:
        meta.setdefault(f, "")

    # list of missing fields
    missing = [f for f in FIELDS if not meta[f].strip()]
    return meta, missing

# —————————————————————————————————————
# 4) MAIN: build one row per PDF
# —————————————————————————————————————

def main():
    pattern = os.path.join(SOURCE_DIR, "*.pdf")
    pdfs = sorted(glob.glob(pattern))
    total = len(pdfs)
    if total == 0:
        logger.error(f"No PDFs found in '{SOURCE_DIR}'")
        return

    records = []
    errors  = []

    for idx, pdf in enumerate(pdfs, start=1):
        name = os.path.basename(pdf)
        logger.info(f"🔍 ({idx}/{total}) {name}")
        try:
            meta, missing = parse_certificate(pdf, idx)
        except Exception as e:
            errors.append(f"{name}: EXCEPTION: {e}")
            # on exception, record only Sl.No and Filename
            meta = {"Sl. No": idx, "Filename": name}
            for f in FIELDS:
                meta.setdefault(f, "")
            records.append(meta)
            continue

        if missing:
            errors.append(f"{name}: missing {len(missing)} → {missing}")
            # blank out everything except Sl. No & Filename
            blank = {"Sl. No": idx, "Filename": name}
            for f in FIELDS:
                blank.setdefault(f, "")
            records.append(blank)
        else:
            records.append(meta)

    # write to Excel
    df = pd.DataFrame(records, columns=["Sl. No"] + FIELDS)
    df.to_excel(OUT_XLSX, sheet_name=SHEET_NAME, index=False)

    # auto–style with openpyxl
    wb = load_workbook(OUT_XLSX)
    ws = wb[SHEET_NAME]

    # bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # auto–fit every column to its longest cell
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(OUT_XLSX)

    # summary logs
    logger.info(f"✅ Wrote {len(records)} rows to '{OUT_XLSX}'")
    if errors:
        logger.warning("⚠️ Some PDFs had issues:")
        for e in errors:
            logger.warning(" - %s", e)

if __name__ == "__main__":
    main()
