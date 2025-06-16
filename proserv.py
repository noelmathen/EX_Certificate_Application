#!/usr/bin/env python3
import os
import glob
import logging

import fitz           # PyMuPDF
import camelot        # pip install "camelot-py[cv]"
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# —————————————————————————————————————
# CONFIGURATION
# —————————————————————————————————————

SOURCE_DIR   = "Proserv Certificates"
OUT_XLSX      = "Proserv_Certificates.xlsx"
SHEET_NAME = "Proserv Certificates"
EXPECTED_COLS = 18

# —————————————————————————————————————
# LOGGER SETUP
# —————————————————————————————————————

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

error_logs = []
all_dfs    = []

# —————————————————————————————————————
# MAIN EXTRACTION
# —————————————————————————————————————

# 1) Find all PDFs in SOURCE_DIR
pattern = os.path.join(SOURCE_DIR, "*.pdf")
pdfs = sorted(glob.glob(pattern))
if not pdfs:
    logger.error(f"No PDFs found in '{SOURCE_DIR}'")
    raise SystemExit(1)

for pdf in pdfs:
    name = os.path.basename(pdf)
    logger.info(f"🔍 Processing: {name}")
    equipment_id = ""

    # --- 2) Open PDF ---
    try:
        doc = fitz.open(pdf)
    except Exception as e:
        error_logs.append(f"{name}: failed to open PDF ({e})")
        continue

    # --- 3) Extract Equipment ID from page 1 ---
    try:
        lines = doc[0].get_text().splitlines()
        idx = next(i for i, ln in enumerate(lines) if "Equipment ID" in ln)
        equipment_id = lines[idx + 5].strip()
        # logger.info(f"   ✔ Found Equipment ID: {equipment_id}")
    except Exception as e:
        error_logs.append(f"{name}: Equipment ID extraction failed ({e})")
        equipment_id = ""

    # --- 4) Determine last page number ---
    last_page = doc.page_count
    doc.close()

    # --- 5) Extract table via Camelot (lattice → stream) ---
    tables = []
    try:
        tables = camelot.read_pdf(pdf, pages=str(last_page),
                                  flavor="lattice", line_scale=40)
    except Exception as e:
        error_logs.append(f"{name}: camelot lattice error ({e})")

    if not tables:
        try:
            tables = camelot.read_pdf(pdf, pages=str(last_page),
                                      flavor="stream")
        except Exception as e:
            error_logs.append(f"{name}: camelot stream error ({e})")
        if not tables:
            error_logs.append(f"{name}: no table found on last page")
            continue

    # --- 6) Take the first table and copy ---
    try:
        df = tables[0].df.copy()
    except Exception as e:
        error_logs.append(f"{name}: failed to copy table df ({e})")
        continue

    # --- 7) Promote header row if it matches your template ---
    try:
        header = df.iloc[0].tolist()
        if all(col in header for col in ["S-No.", "Equipment Tag", "Pass/Fail"]):
            df.columns = header
            df = df[1:].reset_index(drop=True)
        else:
            # error_logs.append(f"{name}: header row not recognized, using raw columns")
            continue
    except Exception as e:
        error_logs.append(f"{name}: header promotion failed ({e})")

    # --- 8) Column count check ---
    if df.shape[1] < EXPECTED_COLS:
        error_logs.append(f"{name}: table has {df.shape[1]} cols (expected ≥{EXPECTED_COLS})")

    # --- 9) Attach metadata & collect ---
    try:
        df["Filename"]     = name
        df["Equipment ID"] = equipment_id
        all_dfs.append(df)
    except Exception as e:
        error_logs.append(f"{name}: failed to append metadata ({e})")

# —————————————————————————————————————
# CONCAT & CLEANUP
# —————————————————————————————————————

if not all_dfs:
    logger.error("❌ No tables extracted from any PDF.")
    for e in error_logs:
        logger.error(" - %s", e)
    raise SystemExit(1)

final = pd.concat(all_dfs, ignore_index=True)

# Define your 18 “data” columns
fixed_cols = [
    "S-No.", "Equipment Tag #", "Equipment Description", "Manufacturer", "Model",
    "Circuit ID", "Area of Classification", "IP", "Protection Method", "Gas Group",
    "T-Rating", "Serial Number", "Certifying Authority", "Certificate No.",
    "Grade of Inspection", "Inspection Date", "Expiry Date", "Pass/Fail"
]

# 10) Trim any extras, rename first 18 columns
final = final.iloc[:, :len(fixed_cols)]
final.columns = fixed_cols

# 11) Rebuild Filename & Equipment ID in order
filenames, eq_ids = [], []
for df_ in all_dfs:
    n = len(df_)
    filenames += [df_["Filename"].iloc[0]] * n
    eq_ids     += [df_["Equipment ID"].iloc[0]] * n

final["Filename"]     = filenames[:len(final)]
final["Equipment ID"] = eq_ids[:len(final)]

# 12) Keep only numeric S-No. rows
mask = final["S-No."].astype(str).str.match(r'^\d+$')
final = final[mask].reset_index(drop=True)

# —————————————————————————————————————
# WRITE TO EXCEL
# —————————————————————————————————————

try:
    final.to_excel(OUT_XLSX, sheet_name="Sheet 2", index=False)

    wb = load_workbook(OUT_XLSX)
    ws = wb["Sheet 2"]

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(OUT_XLSX)
    logger.info(f"✅ Successfully wrote {len(final)} rows to '{OUT_XLSX}'")

except Exception as e:
    error_logs.append(f"Failed to write Excel file ({e})")

# Attempt to clear Camelot tempdirs
try:
    from camelot.utils import _tempdir_manager
    _tempdir_manager._tempdirs.clear()
except:
    pass

# —————————————————————————————————————
# SUMMARY
# —————————————————————————————————————

if error_logs:
    logger.warning("⚠️ Issues encountered during processing:")
    for e in error_logs:
        logger.warning(" - %s", e)
