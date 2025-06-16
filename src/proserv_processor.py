#src/proserv_processor.py
#!/usr/bin/env python3
"""
Proserv Certificate Processor Module
Handles processing of Proserv certificates into Excel format
"""

import os, tempfile
import glob
import logging
import fitz
import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import time
import errno
import functools

class ProservProcessor:
    """Proserv certificate processor"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.setup_robust_rmtree()
        
        # Configuration
        self.SOURCE_DIR = "Proserv Certificates"
        self.OUT_XLSX = "Proserv_Certificates.xlsx"
        self.SHEET_NAME = "Proserv Certificates"
        self.EXPECTED_COLS = 18
        
        # Header detection tokens
        self.TOKENS = ["S-No", "Equipment Tag", "Pass/Fail"]
        
        # Fixed column names
        self.FIXED_COLS = [
            "S-No.", "Equipment Tag #", "Equipment Description", "Manufacturer",
            "Model", "Circuit ID", "Area of Classification", "IP", "Protection Method",
            "Gas Group", "T-Rating", "Serial Number", "Certifying Authority",
            "Certificate No.", "Grade of Inspection", "Inspection Date",
            "Expiry Date", "Pass/Fail"
        ]
        
        # Storage for results
        self.dataframes = []
        self.errors = []
        
    def setup_robust_rmtree(self):
        """Setup robust rmtree to handle Windows file locking issues"""
        _orig_rmtree = shutil.rmtree
        
        def _robust_rmtree(path, *args, **kwargs):
            try:
                _orig_rmtree(path, *args, **kwargs)
            except PermissionError as e:
                if e.errno != errno.EACCES:
                    raise
                time.sleep(0.2)
                try:
                    _orig_rmtree(path, *args, **kwargs)
                except PermissionError:
                    pass  # Still locked, just leave the directory
                    
        shutil.rmtree = _robust_rmtree
        
    def process(self):
        """Main processing function"""
        self.logger.info("Starting Proserv certificate processing...")
        
        # Find PDF files
        pdf_files = sorted(glob.glob(os.path.join(self.SOURCE_DIR, "*.pdf")))
        if not pdf_files:
            self.logger.error(f"No PDFs found in '{self.SOURCE_DIR}'")
            return
            
        self.logger.info(f"Found {len(pdf_files)} Proserv certificates to process")
        
        # Process each PDF
        for pdf_file in pdf_files:
            self._process_single_pdf(pdf_file)
            
        # Generate final Excel file
        self._generate_excel()
        
        # Log summary
        self._log_summary()
        
    def _process_single_pdf(self, pdf_path):
        """Process a single PDF file"""
        filename = os.path.basename(pdf_path)
        self.logger.info(f"🔍 Processing: {filename}")
        
        try:
            # Extract equipment ID from first page
            equipment_id = self._extract_equipment_id(pdf_path)
            
            # Extract table from last page
            dataframe = self._extract_table_data(pdf_path)
            
            if dataframe is not None:
                # Add metadata columns
                dataframe["Filename"] = filename
                dataframe["Equipment ID"] = equipment_id
                self.dataframes.append(dataframe)
                self.logger.info(f"✅ Successfully processed: {filename}")
            else:
                self.errors.append(f"{filename}: Could not extract table data")
                
        except Exception as e:
            self.errors.append(f"{filename}: Exception - {e}")
            self.logger.error(f"❌ Error processing {filename}: {e}")
            
    def _extract_equipment_id(self, pdf_path):
        """Extract equipment ID from the first page"""
        try:
            doc = fitz.open(pdf_path)
            lines = doc[0].get_text().splitlines()
            
            # Find "Equipment ID" line and get the ID from 5 lines below
            for i, line in enumerate(lines):
                if "Equipment ID" in line:
                    if i + 5 < len(lines):
                        equipment_id = lines[i + 5].strip()
                        doc.close()
                        return equipment_id
                        
            doc.close()
            return ""
            
        except Exception as e:
            self.logger.warning(f"Could not extract equipment ID: {e}")
            return ""
            
    def _extract_table_data(self, pdf_path):
        """Extract table data from the last page via a decompressed temp‐PDF + Camelot."""
        temp_pdf = None
        try:
            # 1) Open & save an uncompressed copy
            doc = fitz.open(pdf_path)
            last_page = doc.page_count
            # create a temp file
            tmpf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            temp_pdf = tmpf.name
            tmpf.close()
            doc.save(temp_pdf)
            doc.close()

            # 2) Try lattice (if Ghostscript is on PATH) then stream (no line_scale)
            tables = None
            for flavor in ("lattice", "stream"):
                try:
                    if flavor == "lattice":
                        tables = camelot.read_pdf(
                            temp_pdf,
                            pages=str(last_page),
                            flavor="lattice",
                            line_scale=40
                        )
                    else:
                        tables = camelot.read_pdf(
                            temp_pdf,
                            pages=str(last_page),
                            flavor="stream"
                        )
                    if tables and len(tables) > 0:
                        break
                except Exception as e:
                    self.logger.debug(f"Camelot {flavor} failed: {e}")

            if not tables or len(tables) == 0:
                self.logger.warning(f"No tables found in {os.path.basename(pdf_path)}")
                return None

            # 3) Process the first table exactly as before
            df = tables[0].df.copy()
            df = self._process_headers(df)
            if df.shape[1] < self.EXPECTED_COLS:
                self.logger.warning(
                    f"{os.path.basename(pdf_path)}: Only {df.shape[1]} columns "
                    f"(expected {self.EXPECTED_COLS})"
                )
                return None

            df = df.iloc[:, :self.EXPECTED_COLS]
            df.columns = self.FIXED_COLS
            df = df[df["S-No."].astype(str).str.match(r'^\d+$')].reset_index(drop=True)
            return df

        except Exception as e:
            self.logger.error(f"Table extraction failed: {e}")
            return None

        finally:
            # clean up the temp PDF
            if temp_pdf and os.path.exists(temp_pdf):
                try:
                    os.remove(temp_pdf)
                except Exception:
                    pass
            
    def _process_headers(self, df):
        """Process and detect headers in the dataframe"""
        def has_header_tokens(row):
            """Check if a row contains header identification tokens"""
            cells = [str(cell).lower() for cell in row]
            return all(any(token.lower() in cell for cell in cells) for token in self.TOKENS)
            
        # Look for single-row header
        header_idx = None
        for i, row in df.iterrows():
            if has_header_tokens(row):
                header_idx = i
                break
                
        # If no single-row header found, try merging first two rows
        if header_idx is None and len(df) > 1:
            merged_row = df.iloc[0].astype(str) + " " + df.iloc[1].astype(str)
            if has_header_tokens(merged_row):
                # Use merged header
                df.columns = merged_row.tolist()
                df = df[2:].reset_index(drop=True)
                return df
                
        # Use single-row header
        if header_idx is not None:
            df.columns = df.iloc[header_idx].tolist()
            df = df[header_idx + 1:].reset_index(drop=True)
            
        return df
        
    def _generate_excel(self):
        """Generate the final Excel file"""
        if not self.dataframes:
            self.logger.error("No data to write to Excel")
            return
            
        try:
            # Combine all dataframes
            final_df = pd.concat(self.dataframes, ignore_index=True)
            
            # Add metadata columns at the end
            final_df.insert(len(self.FIXED_COLS), "Filename", final_df.pop("Filename"))
            final_df.insert(len(self.FIXED_COLS) + 1, "Equipment ID", final_df.pop("Equipment ID"))
            
            # Write to Excel
            final_df.to_excel(self.OUT_XLSX, sheet_name=self.SHEET_NAME, index=False)
            
            # Format Excel file
            self._format_excel_file()
            
            self.logger.info(f"✅ Wrote {len(final_df)} rows to '{self.OUT_XLSX}'")
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel file: {e}")
            raise
            
    def _format_excel_file(self):
        """Format the Excel file with proper styling"""
        try:
            wb = load_workbook(self.OUT_XLSX)
            ws = wb[self.SHEET_NAME]
            
            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = max(
                    len(str(cell.value)) for cell in column if cell.value
                ) + 2
                ws.column_dimensions[column[0].column_letter].width = max_length
                
            wb.save(self.OUT_XLSX)
            
        except Exception as e:
            self.logger.warning(f"Could not format Excel file: {e}")
            
    def _log_summary(self):
        """Log processing summary"""
        total_processed = len(self.dataframes)
        total_errors = len(self.errors)
        
        self.logger.info("📊 Proserv Processing Summary:")
        self.logger.info(f"   • Successfully processed: {total_processed}")
        self.logger.info(f"   • Errors encountered: {total_errors}")
        
        if self.errors:
            self.logger.warning("⚠️ Processing issues:")
            for error in self.errors:
                self.logger.warning(f" - {error}")